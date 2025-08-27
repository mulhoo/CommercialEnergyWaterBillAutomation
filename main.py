#!/usr/bin/env python3
"""
Water Bill PDF Processor - Clean Version
Processes batch PDFs, renames files, and generates Excel reports
"""

import os
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Optional
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from dataclasses import dataclass
import re
from calendar import month_date

# PDF and Excel processing
try:
    import pdfplumber
    from openpyxl import load_workbook
    from openpyxl import numbers
    import pytesseract
    from pdf2image import convert_from_path
except ImportError as e:
    raise SystemExit(
        f"{e}\nInstall with:\n  pip install pdfplumber openpyxl pytesseract pdf2image"
    )

# Folder Structure
BASE_DIR = Path("Reports & Bills")
REPORTS_ROOT = BASE_DIR / "Reports"
BILLS_ROOT = BASE_DIR / "Bills"

REPORTS_DIRS = {
    "North Marin": REPORTS_ROOT / "North Marin",
    "Marin Water": REPORTS_ROOT / "Marin Water",
}

BILLS_DIRS = {
    "North Marin": BILLS_ROOT / "North Marin",
    "Marin Water": BILLS_ROOT / "Marin Water",
}

def month_year_folder(bill_date_str: str) -> str:
    """
    bill_date_str is expected as %m/%d/%Y (e.g., 09/15/2025).
    Falls back to current month/year if parsing fails.
    """
    try:
        dt = datetime.strptime(bill_date_str, "%m/%d/%Y")
    except Exception:
        dt = datetime.now()
    return f"{month_name[dt.month]} {dt.year}"


@dataclass
class BillData:
    """Data structure for bill information"""
    account_number: str
    bill_date: str
    due_date: str
    total_due: float
    service_address: str
    current_usage_gallons: int
    service_period: str
    district: str
    original_filename: str

class BillExtractor:
    """Extract data from water bill PDFs"""

    def __init__(self):
        self.logger = self._setup_logging()

    def _setup_logging(self):
        import logging
        logging.basicConfig(level=logging.INFO)
        return logging.getLogger('BillExtractor')

    def extract_data(self, pdf_path: str, district: str) -> Optional[BillData]:
        """Extract data from PDF based on district"""
        if district == "North Marin":
            return self._extract_nmwd_data(pdf_path)
        elif district == "Marin Water":
            return self._extract_mmwd_data(pdf_path)
        else:
            self.logger.error(f"Unknown district: {district}")
            return None

    def _extract_nmwd_data(self, pdf_path: str) -> Optional[BillData]:
        """Extract data from North Marin Water District bill"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()

                if not text or "NORTH MARIN" not in text.upper():
                    text = self._ocr_extract(pdf_path)

                if not text:
                    return None

                # Extract key fields with flexible patterns
                account_number = (
                    self._extract_pattern(text, r'ACCOUNT(?:/CUSTOMER)? NUMBER[:\s]*([A-Z0-9\-]{6,})') or
                    self._extract_pattern(text, r'Customer Number[:\s]*([A-Z0-9\-]{6,})')
                )

                bill_date = self._extract_pattern(text, r'(\d{2}/\d{2}/\d{4})')

                due_date = "Upon Receipt" if "Upon Receipt" in text else \
                          self._extract_pattern(text, r'DUE DATE[^$]*(\d{2}/\d{2}/\d{4})')

                total_due = self._extract_currency(text, r'TOTAL DUE.*?\$?([\d,]+\.?\d*)') or \
                           self._extract_currency(text, r'TOTAL (?:AMOUNT )?DUE.*?\$?\(?([\d,]+\.?\d*)\)?')

                service_address = self._extract_pattern(text, r'SERVICE ADDRESS.*?(\d+[^,\n]*)')

                # Usage data
                current_usage = self._extract_number(text, r'CURRENT PERIOD:?\s*(\d+(?:,\d+)?)') or \
                              self._extract_number(text, r'(\d+)\s+GAL') or 0

                # Extract service period
                period_match = re.search(r'(\d+/\d+/\d{4})\s*-\s*(\d+/\d+/\d{4})', text)
                service_period = period_match.group(0) if period_match else ""

                if not account_number or total_due is None:
                    return None

                return BillData(
                    account_number=account_number,
                    bill_date=bill_date or '',
                    due_date=due_date or "Upon Receipt",
                    total_due=total_due,
                    service_address=service_address or '',
                    current_usage_gallons=current_usage,
                    service_period=service_period,
                    district="North Marin",
                    original_filename=os.path.basename(pdf_path)
                )

        except Exception as e:
            self.logger.error(f"Failed to extract NMWD data from {pdf_path}: {e}")
            return None

    def _extract_mmwd_data(self, pdf_path: str) -> Optional[BillData]:
        """Extract data from Marin Municipal Water District bill"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()

                if not text or "MARIN WATER" not in text.upper():
                    text = self._ocr_extract(pdf_path)

                if not text:
                    return None

                # Extract key fields - MMWD format
                account_number = self._extract_pattern(text, r'Customer Number:?\s*(\d+)')
                bill_date = self._extract_pattern(text, r'Billing Date:?\s*(\d{2}/\d{2}/\d{4})')
                due_date = self._extract_pattern(text, r'Current Charges Due By:?\s*(\d{2}/\d{2}/\d{4})') or "Upon Receipt"
                total_due = self._extract_currency(text, r'TOTAL DUE:?\s*\$?([\d,]+\.?\d*)')
                service_address = self._extract_pattern(text, r'Service Address:?\s+(.+?)(?=\n)')

                # Usage data - MMWD format (units to gallons conversion)
                current_units = (
                    self._extract_number(text, r'Water Use\s+Units\*?\s+(\d+)') or
                    self._extract_number(text, r'Units\s*:\s*(\d+)') or 0
                )
                current_usage_gallons = current_units * 748 # 1 unit = 748 gallons

                # Extract service period
                period_match = re.search(r'(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})', text)
                service_period = period_match.group(0) if period_match else ""

                if not account_number or total_due is None:
                    return None

                return BillData(
                    account_number=account_number,
                    bill_date=bill_date or '',
                    due_date=due_date,
                    total_due=total_due,
                    service_address=service_address or '',
                    current_usage_gallons=current_usage_gallons,
                    service_period=service_period,
                    district="Marin Water",
                    original_filename=os.path.basename(pdf_path)
                )

        except Exception as e:
            self.logger.error(f"Failed to extract MMWD data from {pdf_path}: {e}")
            return None

    def _ocr_extract(self, pdf_path: str) -> Optional[str]:
        """Extract text using OCR for scanned PDFs"""
        try:
            images = convert_from_path(pdf_path, dpi=300)
            text = ""
            for image in images:
                try:
                  text += pytesseract.image_to_string(image, config='--psm 6') + "\n"
                except Exception:
                  text += pytesseract.image_to_string(image, config='--psm 4') + "\n"
            return text
        except Exception as e:
            self.logger.error(f"OCR extraction failed: {e}")
            return None

    def _extract_pattern(self, text: str, pattern: str) -> Optional[str]:
        """Extract first match of regex pattern"""
        match = re.search(pattern, text, re.IGNORECASE)
        return match.group(1).strip() if match else None

    def _extract_currency(self, text, pattern):
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if not m: return None
        s = m.group(1)
        s = s.replace(',', '').replace('$','').strip()
        neg = False
        if s.startswith('(') and s.endswith(')'):
            neg, s = True, s[1:-1]
        try:
            val = float(s)
            return -val if neg else val
        except ValueError:
            return None

    def _extract_number(self, text: str, pattern: str) -> Optional[int]:
        """Extract number and convert to int"""
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value_str = match.group(1).replace(',', '')
            try:
                return int(value_str)
            except ValueError:
                return None
        return None

class FileRenamer:
    """Handle PDF file renaming according to specifications"""

    def generate_filename(self, bill_data: BillData) -> str:
        """Generate new filename based on district and data"""
        try:
            date_obj = datetime.strptime(bill_data.bill_date, "%m/%d/%Y")
            date_str = date_obj.strftime("%y%m%d")
        except:
            date_str = "000000"

        if bill_data.district == "North Marin":
            filename = f"{date_str} Account #{bill_data.account_number}.pdf"
        elif bill_data.district == "Marin Water":
            filename = f"{date_str} MMWD {bill_data.account_number}.pdf"
        else:
            filename = f"{date_str} {bill_data.account_number}.pdf"

        # Clean filename of invalid characters
        filename = "".join(c for c in filename if c.isalnum() or c in " #.-_")
        return filename

    def rename_file(self, original_path: str, new_filename: str, output_dir: str) -> str:
        """Rename and move file to output directory"""
        output_path = Path(output_dir) / new_filename
        shutil.copy2(original_path, output_path)
        return str(output_path)

class ExcelProcessor:
    """Process Excel templates and populate with bill data"""

    def __init__(self):
        self.templates = {
            "North Marin": "BioMarin Pharmaceutical Inc. Account Allocation - North Marin Water - Template for Olivia.xlsx",
            "Marin Water": "BioMarin Pharmaceutical Inc. Account Allocation - Marin Municipal Water District - Template.xlsx"
        }

        self.district_config = {
            "North Marin": {
                "vendor_id": "300011",
                "supplier_name": "North Marin Water District"
            },
            "Marin Water": {
                "vendor_id": "309438",
                "supplier_name": "Marin Municipal Water District"
            }
        }

    def generate_excel_report(self, bills: List[BillData], district: str) -> Optional[str]:
        """Generate Excel report for the specified district"""
        if not bills:
            return None

        try:
            template_path = self.templates[district]

            if not os.path.exists(template_path):
                print(f"Template not found: {template_path}")
                return None

            # Load the template
            workbook = load_workbook(template_path)
            worksheet = workbook.active

            # Clear existing data (rows 9 and below)
            max_row = worksheet.max_row
            for row in range(9, max_row + 1):
                for col in range(1, 10):
                    worksheet.cell(row=row, column=col).value = None

            # Populate with new data
            config = self.district_config[district]
            if worksheet.max_row > 8:
                worksheet.delete_rows(9, worksheet.max_row - 8)

            for i, bill in enumerate(bills, start=9):
                # A: Date of Invoice
                try:
                    dt = datetime.strptime(bill.bill_date, "%m/%d/%Y")
                    c = worksheet.cell(row=i, column=1, value=dt)
                    c.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                except:
                    worksheet.cell(row=i, column=1).value = bill.bill_date

                # B: Supply Address
                worksheet.cell(row=i, column=2).value = bill.service_address

                # C: Service Period
                worksheet.cell(row=i, column=3).value = bill.service_period

                # D: Commodity
                worksheet.cell(row=i, column=4).value = "Water"

                # E: GL Code
                worksheet.cell(row=i, column=5).value = "105-000-60035-803-0000"

                # F: Vendor ID
                worksheet.cell(row=i, column=6).value = config["vendor_id"]

                # G: Supplier
                worksheet.cell(row=i, column=7).value = config["supplier_name"]

                # H: Account Number
                worksheet.cell(row=i, column=8).value = bill.account_number

                # I: Current Charges
                c = worksheet.cell(row=i, column=9, value=bill.total_due)
                c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            # Save the populated file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            district_short = "NMWD" if district == "North Marin" else "MMWD"
            output_filename = f"BioMarin_{district_short}_Report_{timestamp}.xlsx"

            # Reports & Bills/Reports/<District>/
            reports_dir = REPORTS_DIRS[district]
            reports_dir.mkdir(parents=True, exist_ok=True)
            output_path = reports_dir / output_filename

            workbook.save(output_path)
            return str(output_path)

        except Exception as e:
            print(f"Error generating Excel report: {e}")
            return None

class WaterBillProcessorGUI:
    """Main GUI application for water bill processing"""

    def __init__(self, root):
        self.root = root
        self.root.title("Commercial Energy Water Bill PDF Processor")
        self.root.geometry("800x600")

        # Initialize processors
        self.extractor = BillExtractor()
        self.renamer = FileRenamer()
        self.excel_processor = ExcelProcessor()

        # Store selected files
        self.selected_files = []

        # Create output directories
        BASE_DIR.mkdir(exist_ok=True)

        self.setup_gui()

    def setup_gui(self):
        """Setup the GUI components"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Title
        title_label = ttk.Label(main_frame, text="Water Bill PDF Processor",
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # District selection
        district_frame = ttk.LabelFrame(main_frame, text="Select District", padding="10")
        district_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        self.district_var = tk.StringVar(value="North Marin")
        ttk.Radiobutton(district_frame, text="North Marin Water District",
                       variable=self.district_var, value="North Marin").grid(row=0, column=0, sticky=tk.W, padx=(0, 50))
        ttk.Radiobutton(district_frame, text="Marin Municipal Water District",
                       variable=self.district_var, value="Marin Water").grid(row=0, column=1, sticky=tk.W)

        # File selection
        file_frame = ttk.LabelFrame(main_frame, text="File Processing", padding="10")
        file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        # Instructions
        instructions = ttk.Label(file_frame, text="Select specific PDF files to process",
                               font=("Arial", 10))
        instructions.grid(row=0, column=0, columnspan=2, pady=(0, 10))

        ttk.Button(file_frame, text="Select PDF Files",
                  command=self.select_files).grid(row=1, column=0, padx=(0, 10))
        ttk.Button(file_frame, text="Process Files",
                  command=self.process_files).grid(row=1, column=1)

        # Results display
        results_frame = ttk.LabelFrame(main_frame, text="Processing Results", padding="10")
        results_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        # Results treeview
        columns = ("Original File", "Renamed File", "Account", "Amount", "Status")
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=15)

        for col in columns:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=120)

        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=scrollbar.set)

        self.results_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        # Status bar
        self.status_var = tk.StringVar(value="Ready to process files")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

    def select_files(self):
        """Select specific PDF files for processing"""
        files = filedialog.askopenfilenames(
            title="Select Water Bill PDFs",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )

        if files:
            self.selected_files = list(files)
            self.status_var.set(f"Selected {len(files)} PDF files for {self.district_var.get()}")
        else:
            self.selected_files = []

    def process_files(self):
        """Process the selected PDF files"""
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select PDF files first.")
            return

        district = self.district_var.get()
        self.results_tree.delete(*self.results_tree.get_children())

        successful_bills = []

        for file_path in self.selected_files:
            self.status_var.set(f"Processing {os.path.basename(file_path)}...")
            self.root.update()

            # Extract data
            bill_data = self.extractor.extract_data(file_path, district)

            if bill_data:
                try:
                    # Generate new filename
                    new_filename = self.renamer.generate_filename(bill_data)

                    # Decide output dir: Bills/<District>/<Month Year>/
                    month_folder = month_year_folder(bill_data.bill_date)
                    district_bills_dir = BILLS_DIRS[district] / month_folder
                    district_bills_dir.mkdir(parents=True, exist_ok=True)

                    # Rename (copy) into that directory
                    new_path = self.renamer.rename_file(file_path, new_filename, str(district_bills_dir))

                    # Add to results
                    self.results_tree.insert("", "end", values=(
                        bill_data.original_filename,
                        new_filename,
                        bill_data.account_number,
                        f"${bill_data.total_due:,.2f}",
                        "Success"
                    ))

                    successful_bills.append(bill_data)

                except Exception as e:
                    self.results_tree.insert("", "end", values=(
                        os.path.basename(file_path),
                        "Error",
                        "—",
                        "—",
                        f"Rename failed: {str(e)[:30]}"
                    ))
            else:
                self.results_tree.insert("", "end", values=(
                    os.path.basename(file_path),
                    "—",
                    "—",
                    "—",
                    "Extraction failed"
                ))

        # Generate Excel report
        if successful_bills:
            excel_path = self.excel_processor.generate_excel_report(successful_bills, district)
            if excel_path:
                self.status_var.set(
                    f"Processed {len(successful_bills)} files.\nExcel report saved to: {excel_path}"
                )
                messagebox.showinfo(
                    "Success",
                    "Processing complete!\n\n"
                    f"Excel report saved to:\n{excel_path}\n\n"
                    "Bills were copied into:\n"
                    f"Reports & Bills/Bills/{district}/<Month Year>/"
                )
        else:
            self.status_var.set("No files processed successfully.")

def _check_binaries():
    from shutil import which
    missing = []
    if which("tesseract") is None:
        missing.append("Tesseract OCR (brew install tesseract / choco install tesseract)")
    if which("pdftoppm") is None:  # Poppler
        missing.append("Poppler (brew install poppler / choco install poppler)")
    if missing:
        raise RuntimeError("Missing system dependencies:\n- " + "\n- ".join(missing))

def main():
    """Run the application"""
    root = tk.Tk()
    check = _check_binaries()
    app = WaterBillProcessorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()