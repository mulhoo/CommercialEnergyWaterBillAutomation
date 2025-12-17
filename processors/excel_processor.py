"""
Excel template processing functionality - DEBUG VERSION
Replace the content of processors/excel_processor.py with this temporarily
"""

import os
import re
import logging
from typing import List, Optional
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import numbers
except ImportError as e:
    raise SystemExit(f"Missing Excel dependency: {e}")

from models.bill_data import BillData
from config import TEMPLATES, DISTRICT_CONFIG, EXCEL_LAYOUT, REPORTS_DIRS

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Process Excel templates and populate with bill data"""

    @staticmethod
    def _norm_acct(value) -> str:
        """Normalize account strings/numbers to digits-only for comparison."""
        if value is None:
            return ""
        return re.sub(r"\D", "", str(value))

    @staticmethod
    def _is_account_match(bill_account: str, excel_account: str) -> bool:
        """Check if bill account matches excel account (handles partial matches)"""
        if not bill_account or not excel_account:
            return False

        bill_norm = re.sub(r"\D", "", str(bill_account))
        excel_norm = re.sub(r"\D", "", str(excel_account))

        if bill_norm == excel_norm:
            return True

        if bill_norm in excel_norm:
            return True

        if excel_norm in bill_norm:
            return True

        return False

    @staticmethod
    def _is_blank(cell) -> bool:
        """Check if a cell is blank or contains only whitespace"""
        cell_value = cell.value
        return cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == "")

    def generate_excel_report(self, bills: List[BillData], district: str) -> Optional[str]:
        """Fill existing template rows by matching Account Number (col H). Do NOT add/delete rows."""
        if not bills:
            logger.error("ERROR: No bills provided to generate_excel_report")
            return None

        self.last_unmatched = []

        try:
            template_path = TEMPLATES[district]
            logger.info("="*60)
            logger.info(f"EXCEL GENERATION DEBUG - {district}")
            logger.info("="*60)
            logger.info(f"Template path: {template_path}")
            logger.info(f"Current directory: {os.getcwd()}")
            logger.info(f"Template exists: {os.path.exists(template_path)}")
            
            # List Excel files in current directory
            excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
            logger.info(f"Excel files in current directory: {len(excel_files)}")
            for f in excel_files:
                logger.info(f"   - {f}")
            
            if not os.path.exists(template_path):
                logger.error(f"ERROR: Template not found at {template_path}")
                return None

            logger.info(f"Loading workbook...")
            workbook = load_workbook(template_path)
            worksheet = workbook.active
            logger.info(f"Workbook loaded - Sheet: '{worksheet.title}'")
            
            config = DISTRICT_CONFIG[district]
            start_row = EXCEL_LAYOUT["start_row"]
            account_col = EXCEL_LAYOUT["account_col"]
            logger.info(f"Start row: {start_row}, Account column: {account_col}")

            excel_accounts = {}

            # Read accounts from template
            logger.info(f"\nScanning for accounts in template...")
            for row_num in range(start_row, min(start_row + 50, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_num, column=account_col).value
                if cell_value:
                    excel_account = str(cell_value).strip()
                    excel_accounts[row_num] = excel_account

            logger.info(f"Found {len(excel_accounts)} accounts in template")
            if excel_accounts:
                logger.info(f"Sample accounts (first 5):")
                for row, acct in list(excel_accounts.items())[:5]:
                    logger.info(f"   Row {row}: {acct}")
            else:
                logger.warning(f"WARNING: No accounts found!")
                logger.info(f"Checking cells around expected location:")
                for r in range(max(1, start_row-2), start_row+5):
                    row_data = []
                    for c in range(max(1, account_col-2), account_col+3):
                        cell_val = worksheet.cell(row=r, column=c).value
                        row_data.append(f"({r},{c})={cell_val}")
                    logger.info(f"   {' | '.join(row_data)}")

            # Process each bill
            logger.info(f"\nProcessing {len(bills)} bills...")
            matched_count = 0
            
            for i, bill in enumerate(bills, 1):
                logger.info(f"\n  Bill {i}/{len(bills)}: Account {bill.account_number}")
                target_row = None

                # Try to find blank matching row first
                for row_num, excel_account in excel_accounts.items():
                    if self._is_account_match(bill.account_number, excel_account):
                        if self._is_blank(worksheet.cell(row=row_num, column=9)):
                            target_row = row_num
                            logger.info(f"    Found BLANK row {row_num} (Excel account: {excel_account})")
                            break

                # If no blank row, use any matching row
                if target_row is None:
                    for row_num, excel_account in excel_accounts.items():
                        if self._is_account_match(bill.account_number, excel_account):
                            target_row = row_num
                            logger.info(f"    Found OCCUPIED row {row_num} (Excel account: {excel_account})")
                            break

                if target_row:
                    logger.info(f"    Populating row {target_row}")
                    self._populate_row(worksheet, target_row, bill, config)
                    matched_count += 1
                else:
                    logger.warning(f"    NO MATCH FOUND for account {bill.account_number}")
                    self.last_unmatched.append((bill.account_number, bill.original_filename))

            logger.info(f"\nSummary: {matched_count}/{len(bills)} bills matched")
            if self.last_unmatched:
                logger.warning(f"Unmatched accounts:")
                for acct, filename in self.last_unmatched:
                    logger.warning(f"   - {acct} ({filename})")

            # Generate output path
            output_path = self._generate_output_path(district, bills)
            logger.info(f"\nOutput path: {output_path}")
            logger.info(f"Output directory: {output_path.parent}")
            logger.info(f"Directory exists: {output_path.parent.exists()}")
            
            # Create directory if needed
            try:
                output_path.parent.mkdir(parents=True, exist_ok=True)
                logger.info(f"Directory ready")
            except Exception as e:
                logger.error(f"ERROR creating directory: {e}")
                return None

            # Check if file is already open
            if output_path.exists():
                logger.info(f"File already exists, will overwrite")
                try:
                    # Test if we can write to it
                    with open(output_path, 'a'):
                        pass
                    logger.info(f"File is not locked")
                except PermissionError:
                    logger.error(f"ERROR: File is open in Excel or locked!")
                    logger.error(f"   Please close the file and try again.")
                    return None

            # Save the workbook
            logger.info(f"Saving workbook...")
            try:
                workbook.save(output_path)
                logger.info(f"SUCCESS: Excel saved to {output_path}")
                logger.info("="*60)
                return str(output_path)
            except PermissionError as e:
                logger.error(f"ERROR: Permission denied - {e}")
                logger.error(f"   File may be open in Excel!")
                return None
            except Exception as e:
                logger.error(f"ERROR saving workbook: {e}")
                return None

        except Exception as e:
            import traceback
            logger.error(f"\nFATAL ERROR in generate_excel_report:")
            logger.error(f"   {type(e).__name__}: {e}")
            logger.error(f"\nFull traceback:")
            logger.error(traceback.format_exc())
            logger.info("="*60)
            return None

    def _populate_row(self, worksheet, row_num: int, bill: BillData, config: dict):
        """Populate a single row with bill data"""

        date_cell = worksheet.cell(row=row_num, column=1)
        if self._is_blank(date_cell):
            try:
                invoice_date = datetime.strptime(bill.bill_date, "%m/%d/%Y")
                date_cell.value = invoice_date
                date_cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
            except Exception:
                date_cell.value = bill.bill_date

        address_cell = worksheet.cell(row=row_num, column=2)
        if self._is_blank(address_cell):
            address_cell.value = bill.service_address

        period_cell = worksheet.cell(row=row_num, column=3)
        if self._is_blank(period_cell):
            period_cell.value = bill.service_period

        commodity_cell = worksheet.cell(row=row_num, column=4)
        if self._is_blank(commodity_cell):
            commodity_cell.value = "Water"

        gl_cell = worksheet.cell(row=row_num, column=5)
        if self._is_blank(gl_cell):
            gl_cell.value = "105-000-60035-803-0000"

        vendor_cell = worksheet.cell(row=row_num, column=6)
        if self._is_blank(vendor_cell):
            vendor_cell.value = config["vendor_id"]

        supplier_cell = worksheet.cell(row=row_num, column=7)
        if self._is_blank(supplier_cell):
            supplier_cell.value = config["supplier_name"]

        account_cell = worksheet.cell(row=row_num, column=8)
        if self._is_blank(account_cell):
            account_cell.value = bill.account_number

        charges_cell = worksheet.cell(row=row_num, column=9)
        if self._is_blank(charges_cell):
            charges_cell.value = bill.total_due
            charges_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        usage_cell = worksheet.cell(row=row_num, column=10)
        if self._is_blank(usage_cell):
            try:
                gallons = int(bill.current_usage_gallons)
            except Exception:
                gallons = bill.current_usage_gallons
            usage_cell.value = gallons
            usage_cell.number_format = '#,##0'

    def _generate_output_path(self, district: str, bills: List[BillData]) -> Path:
        """Generate output path for the Excel report - both districts save to same Pending Invoice folder"""
        reports_dir = REPORTS_DIRS[district]
        
        logger.info(f"Reports directory: {reports_dir}")
        
        try:
            reports_dir.mkdir(parents=True, exist_ok=True)
            logger.info(f"✓ Reports directory exists/created")
        except Exception as e:
            logger.error(f"❌ Cannot create reports directory: {e}")

        # Different filenames for each district
        if district == "North Marin":
            output_filename = "BioMarin Pharmaceutical Inc. Account Allocation - North Marin Water.xlsx"
        else:  # Marin Municipal
            output_filename = "BioMarin Pharmaceutical Inc. Account Allocation - Marin Municipal Water District.xlsx"
        
        output_path = reports_dir / output_filename
        
        logger.info(f"Full output path: {output_path}")
        return output_path